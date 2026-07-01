import io
import datetime
from decimal import Decimal
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from accounts.decorators import cashier_or_admin_required
from .models import SalaryRecord
from employees.models import Employee

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from xhtml2pdf import pisa
    HAS_XHTML2PDF = True
except ImportError:
    HAS_XHTML2PDF = False


@login_required
@cashier_or_admin_required
def export_excel(request):
    """Export salary records to Excel file."""
    if not HAS_OPENPYXL:
        return HttpResponse('openpyxl is not installed. Run: pip install openpyxl', status=500)

    month = request.GET.get('month', '')
    year = request.GET.get('year', '')
    records = SalaryRecord.objects.all().select_related('employee')

    if month:
        records = records.filter(salary_month__icontains=month)
    if year:
        records = records.filter(month_year__year=int(year))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Salary Report'

    # Styles
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    currency_format = '₹#,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells('A1:U1')
    title_cell = ws['A1']
    title_cell.value = f'Employee Salary Payment Register'
    title_cell.font = Font(name='Calibri', bold=True, size=16, color='1B2A4A')
    title_cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:U2')
    subtitle = ws['A2']
    period = f'{month} {year}' if month and year else 'All Records'
    subtitle.value = f'Period: {period} | Generated: {datetime.datetime.now().strftime("%d-%m-%Y %H:%M")}'
    subtitle.font = Font(name='Calibri', size=10, italic=True)
    subtitle.alignment = Alignment(horizontal='center')

    # Headers
    headers = [
        'Sl.No', 'Employee ID', 'Employee Name', 'Designation', 'Department',
        'Date of Joining', 'Salary Month', 'Monthly Salary', 'Working Days',
        'Present Days', 'Absent Days', 'Salary Earned', 'Deduction',
        'Deduction Remarks', 'Net Salary', 'Paid Amount', 'Payment Date',
        'Balance Salary', 'Balance Paid', 'Balance Paid Date', 'Total Paid',
        'Outstanding', 'Payment Status', 'Remarks'
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    currency_cols = [8, 12, 13, 15, 16, 18, 19, 21, 22]
    status_fills = {
        'PENDING': PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid'),
        'PARTIALLY_PAID': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),
        'PAID': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),
    }

    for row_idx, record in enumerate(records, 5):
        data = [
            row_idx - 4,
            record.employee.employee_id,
            record.employee.name,
            record.employee.designation,
            record.employee.department,
            record.employee.date_of_joining.strftime('%d-%m-%Y'),
            record.salary_month,
            float(record.monthly_salary),
            record.total_working_days,
            record.present_days,
            record.absent_days,
            float(record.salary_earned),
            float(record.deduction_amount),
            record.deduction_remarks,
            float(record.net_salary),
            float(record.paid_amount),
            record.payment_date.strftime('%d-%m-%Y') if record.payment_date else '',
            float(record.balance_salary),
            float(record.balance_paid),
            record.balance_paid_date.strftime('%d-%m-%Y') if record.balance_paid_date else '',
            float(record.total_salary_paid),
            float(record.outstanding_balance),
            record.get_payment_status_display(),
            record.remarks,
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

            if col_idx in currency_cols:
                cell.number_format = currency_format

            if col_idx == 23:  # Status column
                fill = status_fills.get(record.payment_status)
                if fill:
                    cell.fill = fill

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Salary_Report_{month}_{year}.xlsx' if month and year else 'Salary_Report_All.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@cashier_or_admin_required
def export_salary_slip_pdf(request, pk):
    """Generate PDF salary slip for a single salary record."""
    if not HAS_XHTML2PDF:
        return HttpResponse('xhtml2pdf is not installed. Run: pip install xhtml2pdf', status=500)

    record = SalaryRecord.objects.select_related('employee').get(pk=pk)

    html_string = render_to_string('salary/salary_slip_pdf.html', {
        'record': record,
        'company_name': 'Your Company Name',
        'generated_at': datetime.datetime.now(),
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="Salary_Slip_{record.employee.employee_id}_{record.salary_month}.pdf"'
    )

    pisa_status = pisa.CreatePDF(html_string, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)

    return response
